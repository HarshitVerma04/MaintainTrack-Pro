"""
export_parts.py — MaintainTrack Pro
--------------------------------------
Exports parts usage and inventory data to an Excel workbook.

Called from Java via ProcessBuilder (ReportsController.java):
    python scripts/export_parts.py

Output: reports/parts_export_YYYY-MM-DD.xlsx

Sheets:
    1. Summary        — all parts with current stock levels
    2. Issue History  — all issue/return transactions
    3. Cost Per Asset — total parts spend grouped by equipment
    4. Low Stock      — parts at or below minimum quantity

Dependencies:
    pip install openpyxl pandas
"""

import sqlite3
import os
import sys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────
# ── Path resolution ──────────────────────────────────────────────────────
# In production (PyInstaller exe), __file__ is inside a temp dir.
# Java sets MAINTAINTRACK_REPORTS_DIR env var to the AppData reports folder.
# In development, fall back to the local reports/ folder.

def _get_db_path():
    # Check AppData first (production)
    appdata = os.environ.get('APPDATA') or os.path.expanduser('~')
    prod_db = os.path.join(appdata, 'MaintainTrackPro', 'maintaintrack.db')
    if os.path.exists(prod_db):
        return prod_db
    # Development fallback
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data', 'maintaintrack.db')

def _get_reports_dir():
    # Java passes this env var in production
    env_dir = os.environ.get('MAINTAINTRACK_REPORTS_DIR')
    if env_dir:
        return env_dir
    # Development fallback
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'reports')

DB_PATH     = _get_db_path()
REPORTS_DIR = _get_reports_dir()
OUTPUT_FILE = os.path.join(REPORTS_DIR, f"parts_export_{date.today()}.xlsx")

# ── Colours ───────────────────────────────────────────────────────────────
NAVY        = "1B3A6B"
BLUE        = "2E86DE"
LIGHT_BLUE  = "D6E8FA"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F4F6FA"
RED_BG      = "FDECEA"
RED_FG      = "8B1A1A"
GREEN_FG    = "1A7A4A"
AMBER_BG    = "FFF3CD"
AMBER_FG    = "B85C00"


# ══════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════

def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def navy_fill():
    return PatternFill("solid", fgColor=NAVY)

def blue_fill():
    return PatternFill("solid", fgColor=BLUE)

def light_fill():
    return PatternFill("solid", fgColor=LIGHT_BLUE)

def gray_fill():
    return PatternFill("solid", fgColor=LIGHT_GRAY)

def red_fill():
    return PatternFill("solid", fgColor=RED_BG)

def amber_fill():
    return PatternFill("solid", fgColor=AMBER_BG)

def thin_border():
    side = Side(style="thin", color="D0D7E3")
    return Border(left=side, right=side, top=side, bottom=side)

def header_font(white=True):
    return Font(name="Calibri", bold=True, size=11,
                color=WHITE if white else NAVY)

def body_font(bold=False, color=None):
    return Font(name="Calibri", bold=bold, size=10,
                color=color or "1A1A2E")

def center():
    return Alignment(horizontal="center", vertical="center")

def left():
    return Alignment(horizontal="left", vertical="center")


def write_sheet_header(ws, title, subtitle=None):
    """Navy title bar at top of each sheet."""
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
    ws["A1"].fill      = navy_fill()
    ws["A1"].alignment = left()
    ws.row_dimensions[1].height = 28

    if subtitle:
        ws.merge_cells("A2:H2")
        ws["A2"] = subtitle
        ws["A2"].font      = Font(name="Calibri", italic=True, size=10, color="5C6B8A")
        ws["A2"].alignment = left()
        ws.row_dimensions[2].height = 18
        return 3  # next data row
    return 2


def write_col_headers(ws, row, headers):
    """Blue column header row."""
    for col, (label, width) in enumerate(headers, start=1):
        cell           = ws.cell(row=row, column=col, value=label)
        cell.font      = header_font()
        cell.fill      = blue_fill()
        cell.alignment = center()
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 20
    return row + 1


def style_data_row(ws, row_num, num_cols, alt=False, fill=None):
    """Apply alternating row fill and borders."""
    row_fill = fill or (gray_fill() if alt else PatternFill("solid", fgColor=WHITE))
    for col in range(1, num_cols + 1):
        cell        = ws.cell(row=row_num, column=col)
        cell.fill   = row_fill
        cell.border = thin_border()
        if cell.alignment.horizontal not in ("center",):
            cell.alignment = left()


# ══════════════════════════════════════════════════════════════════════════
# SHEET 1 — PARTS SUMMARY
# ══════════════════════════════════════════════════════════════════════════

def build_summary_sheet(ws, conn):
    ws.title = "Parts Summary"
    next_row = write_sheet_header(
        ws, "Parts & Inventory Summary",
        f"Generated: {date.today()}  |  All parts with current stock levels"
    )

    headers = [
        ("ID",            6),
        ("Part Name",    32),
        ("Supplier",     22),
        ("Qty on Hand",  13),
        ("Min Qty",      10),
        ("Unit",          9),
        ("Unit Cost",    13),
        ("Stock Status", 14),
    ]
    data_row = write_col_headers(ws, next_row, headers)

    rows = conn.execute("""
        SELECT p.id, p.name, COALESCE(s.name,'—') AS supplier,
               p.qty_on_hand, p.min_qty, p.unit, p.unit_cost
        FROM PART p
        LEFT JOIN SUPPLIER s ON p.supplier_id = s.id
        ORDER BY p.name;
    """).fetchall()

    for i, r in enumerate(rows):
        low   = r["qty_on_hand"] <= r["min_qty"]
        fill  = red_fill() if low else None
        alt   = i % 2 == 0

        ws.cell(data_row, 1, r["id"])
        ws.cell(data_row, 2, r["name"])
        ws.cell(data_row, 3, r["supplier"])
        ws.cell(data_row, 4, r["qty_on_hand"])
        ws.cell(data_row, 5, r["min_qty"])
        ws.cell(data_row, 6, r["unit"])

        cost_cell       = ws.cell(data_row, 7, r["unit_cost"])
        cost_cell.number_format = '₹#,##0.00'

        status          = ws.cell(data_row, 8,
                                   "⚠ Low Stock" if low else "✓ OK")
        status.font     = body_font(bold=True,
                                     color=RED_FG if low else GREEN_FG)
        status.alignment = center()

        style_data_row(ws, data_row, 8, alt=alt, fill=fill)

        # Qty colour
        qty_cell       = ws.cell(data_row, 4)
        qty_cell.font  = body_font(bold=True,
                                    color=RED_FG if low else GREEN_FG)
        qty_cell.alignment = center()

        data_row += 1

    # Summary totals
    data_row += 1
    ws.cell(data_row, 1, f"Total parts: {len(rows)}")
    ws.cell(data_row, 1).font = body_font(bold=True, color=NAVY)
    low_count = sum(1 for r in rows if r["qty_on_hand"] <= r["min_qty"])
    ws.cell(data_row, 4, f"{low_count} low stock")
    ws.cell(data_row, 4).font = body_font(bold=True, color=RED_FG)


# ══════════════════════════════════════════════════════════════════════════
# SHEET 2 — ISSUE HISTORY
# ══════════════════════════════════════════════════════════════════════════

def build_issue_history_sheet(ws, conn):
    ws.title = "Issue History"
    next_row = write_sheet_header(
        ws, "Parts Issue & Return History",
        "All transactions — issues and returns — sorted newest first"
    )

    headers = [
        ("ID",          6),
        ("Date",       14),
        ("Part",       30),
        ("Equipment",  26),
        ("Qty",         7),
        ("Unit",        8),
        ("Type",       12),
        ("Issued By",  16),
    ]
    data_row = write_col_headers(ws, next_row, headers)

    rows = conn.execute("""
        SELECT ir.id, ir.issued_on, p.name AS part, e.name AS equipment,
               ir.qty, p.unit, ir.type, COALESCE(ir.issued_by,'—') AS issued_by
        FROM ISSUE_RECORD ir
        JOIN PART p      ON ir.part_id      = p.id
        JOIN EQUIPMENT e ON ir.equipment_id = e.id
        ORDER BY ir.issued_on DESC;
    """).fetchall()

    for i, r in enumerate(rows):
        is_return = r["type"] == "return"
        row_fill  = amber_fill() if is_return else None

        ws.cell(data_row, 1, r["id"])
        ws.cell(data_row, 2, r["issued_on"])
        ws.cell(data_row, 3, r["part"])
        ws.cell(data_row, 4, r["equipment"])

        qty_cell       = ws.cell(data_row, 5, r["qty"])
        qty_cell.alignment = center()

        ws.cell(data_row, 6, r["unit"])

        type_cell       = ws.cell(data_row, 7,
                                   "↩ Return" if is_return else "↗ Issue")
        type_cell.font  = body_font(bold=True,
                                     color=AMBER_FG if is_return else GREEN_FG)
        type_cell.alignment = center()

        ws.cell(data_row, 8, r["issued_by"])

        style_data_row(ws, data_row, 8, alt=(i % 2 == 0), fill=row_fill)
        data_row += 1

    data_row += 1
    total_issues  = sum(1 for r in rows if r["type"] == "issue")
    total_returns = sum(1 for r in rows if r["type"] == "return")
    ws.cell(data_row, 1,
            f"Total: {len(rows)} transactions  |  "
            f"{total_issues} issues  |  {total_returns} returns"
            ).font = body_font(bold=True, color=NAVY)


# ══════════════════════════════════════════════════════════════════════════
# SHEET 3 — COST PER ASSET
# ══════════════════════════════════════════════════════════════════════════

def build_cost_sheet(ws, conn):
    ws.title = "Cost Per Asset"
    next_row = write_sheet_header(
        ws, "Parts Cost Per Equipment",
        "Sum of (qty x unit_cost) for all issued parts, grouped by machine"
    )

    headers = [
        ("Equipment",    30),
        ("Location",     20),
        ("Issues",        9),
        ("Total Cost",   16),
        ("Cost / Day",   14),
        ("% of Total",   12),
    ]
    data_row = write_col_headers(ws, next_row, headers)

    rows = conn.execute("""
        SELECT e.name, e.location,
               COUNT(ir.id)                          AS issue_count,
               ROUND(SUM(ir.qty * p.unit_cost), 2)  AS total_cost
        FROM ISSUE_RECORD ir
        JOIN EQUIPMENT e ON ir.equipment_id = e.id
        JOIN PART p      ON ir.part_id      = p.id
        WHERE ir.type = 'issue'
        GROUP BY e.id
        ORDER BY total_cost DESC;
    """).fetchall()

    grand_total = sum(r["total_cost"] for r in rows) or 1

    for i, r in enumerate(rows):
        cost_per_day = r["total_cost"] / 90
        pct          = (r["total_cost"] / grand_total) * 100

        ws.cell(data_row, 1, r["name"])
        ws.cell(data_row, 2, r["location"])

        cnt = ws.cell(data_row, 3, r["issue_count"])
        cnt.alignment = center()

        cost = ws.cell(data_row, 4, r["total_cost"])
        cost.number_format = '₹#,##0.00'
        cost.font = body_font(bold=True)

        cpd = ws.cell(data_row, 5, round(cost_per_day, 2))
        cpd.number_format = '₹#,##0.00'

        pct_cell = ws.cell(data_row, 6, round(pct, 1))
        pct_cell.number_format = '0.0"%"'
        pct_cell.alignment = center()

        style_data_row(ws, data_row, 6, alt=(i % 2 == 0))
        data_row += 1

    # Grand total row
    ws.cell(data_row, 1, "GRAND TOTAL").font = header_font(white=False)
    ws.cell(data_row, 1).fill = light_fill()

    total_cell = ws.cell(data_row, 4, grand_total)
    total_cell.number_format = '₹#,##0.00'
    total_cell.font = Font(name="Calibri", bold=True, size=11, color=NAVY)
    total_cell.fill = light_fill()

    for col in range(1, 7):
        ws.cell(data_row, col).fill   = light_fill()
        ws.cell(data_row, col).border = thin_border()


# ══════════════════════════════════════════════════════════════════════════
# SHEET 4 — LOW STOCK
# ══════════════════════════════════════════════════════════════════════════

def build_low_stock_sheet(ws, conn):
    ws.title = "Low Stock Alerts"
    next_row = write_sheet_header(
        ws, "Low Stock Alert List",
        "Parts at or below minimum quantity — reorder required"
    )

    headers = [
        ("Part Name",    32),
        ("Supplier",     24),
        ("Contact",      20),
        ("Phone",        18),
        ("Qty on Hand",  13),
        ("Min Qty",      10),
        ("Unit",          9),
        ("Unit Cost",    13),
    ]
    data_row = write_col_headers(ws, next_row, headers)

    rows = conn.execute("""
        SELECT p.name, COALESCE(s.name,'—') AS supplier,
               COALESCE(s.contact_name,'—') AS contact,
               COALESCE(s.phone,'—')        AS phone,
               p.qty_on_hand, p.min_qty, p.unit, p.unit_cost
        FROM PART p
        LEFT JOIN SUPPLIER s ON p.supplier_id = s.id
        WHERE p.qty_on_hand <= p.min_qty
        ORDER BY p.qty_on_hand ASC;
    """).fetchall()

    if not rows:
        ws.cell(data_row, 1, "All parts are above minimum stock. No alerts.").font = \
            body_font(bold=True, color=GREEN_FG)
        return

    for i, r in enumerate(rows):
        ws.cell(data_row, 1, r["name"])
        ws.cell(data_row, 2, r["supplier"])
        ws.cell(data_row, 3, r["contact"])
        ws.cell(data_row, 4, r["phone"])

        qty = ws.cell(data_row, 5, r["qty_on_hand"])
        qty.font      = body_font(bold=True, color=RED_FG)
        qty.alignment = center()

        mn = ws.cell(data_row, 6, r["min_qty"])
        mn.alignment  = center()

        ws.cell(data_row, 7, r["unit"])

        cost = ws.cell(data_row, 8, r["unit_cost"])
        cost.number_format = '₹#,##0.00'

        style_data_row(ws, data_row, 8, fill=red_fill())
        data_row += 1

    data_row += 1
    ws.cell(data_row, 1,
            f"{len(rows)} part(s) require reordering."
            ).font = body_font(bold=True, color=RED_FG)


# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════

def main():
    if not os.path.exists(DB_PATH):
        print("[ERROR] Database not found. Run Java app first.")
        sys.exit(1)

    os.makedirs(REPORTS_DIR, exist_ok=True)

    conn = get_connection()

    wb = Workbook()

    # Remove default blank sheet
    wb.remove(wb.active)

    # Build all sheets
    build_summary_sheet(wb.create_sheet(),     conn)
    build_issue_history_sheet(wb.create_sheet(), conn)
    build_cost_sheet(wb.create_sheet(),        conn)
    build_low_stock_sheet(wb.create_sheet(),   conn)

    conn.close()

    wb.save(OUTPUT_FILE)
    print(f"[OK] Excel export saved -> {os.path.abspath(OUTPUT_FILE)}")


if __name__ == '__main__':
    main()
